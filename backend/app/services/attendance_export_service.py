"""
考勤记录 Excel 导出服务（吉利学院学风督查记录表格式，含请假条图片）
"""

from typing import Optional
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from sqlalchemy.orm import Session

from ..config import settings
from ..models import AttendanceRecord, ClassRoster, LeaveSlip
from .period_utils import normalize_period

DEFAULT_COLLEGE = "外国语言与文化学院"

IMG_WIDTH_PX = 100
IMG_HEIGHT_PX = 75
IMG_ROW_HEIGHT = 60

HEADERS = [
    "序号", "学院", "年级专业班级", "时间", "周数", "节次", "教室",
    "课程名称", "授课教师", "辅导员",
    "班级人数", "实到人数", "病公假人数", "事假人数", "迟到人数", "早退人数", "旷课人数", "未到人数",
    "平均到课率", "平均到课率（含病公假）",
    "请假、违纪情况说明",
    "学生请假截图（吉利相伴截图）", "学生请假截图2", "学生请假截图3", "学生请假截图4",
]

COL_WIDTHS = [6, 18, 22, 8, 8, 8, 10, 18, 10, 10, 8, 8, 10, 8, 8, 8, 8, 8, 10, 16, 30, 16, 16, 16, 16]


def _format_export_date(date_str: str) -> str:
    if re.match(r"^\d{4}-\d{2}-\d{2}$", date_str):
        _, m, d = date_str.split("-")
        return f"{int(m)}.{int(d)}"
    return date_str


def _get_college(db: Session, class_name: str) -> str:
    if class_name:
        roster = db.query(ClassRoster).filter(ClassRoster.class_name == class_name).first()
        if roster and roster.major:
            return roster.major
    return DEFAULT_COLLEGE


def _resolve_upload_path(image_path: str) -> Optional[Path]:
    if not image_path:
        return None
    rel = image_path.lstrip("/").replace("uploads/", "", 1)
    abs_path = Path(settings.UPLOAD_DIR) / rel
    if abs_path.exists():
        return abs_path
    # 兼容历史路径写法
    alt = Path(settings.UPLOAD_DIR) / Path(image_path).name
    if alt.exists():
        return alt
    return None


def _embed_image(ws, image_path: str, cell_ref: str, row_idx: int):
    abs_path = _resolve_upload_path(image_path)
    if abs_path is None:
        ws[cell_ref].value = "（图片文件不存在）"
        return
    try:
        pil_img = PILImage.open(abs_path)
        if pil_img.mode in ("RGBA", "P", "LA"):
            pil_img = pil_img.convert("RGB")
        buf = BytesIO()
        pil_img.save(buf, format="JPEG", quality=85)
        buf.seek(0)
        img = XlImage(buf)
        img.width = IMG_WIDTH_PX
        img.height = IMG_HEIGHT_PX
        ws.row_dimensions[row_idx].height = max(ws.row_dimensions[row_idx].height or 0, IMG_ROW_HEIGHT)
        ws.add_image(img, cell_ref)
    except Exception:
        ws[cell_ref].value = "（图片加载失败）"


def export_attendance_excel(
    records: list[AttendanceRecord],
    leave_slips_map: dict[int, list[LeaveSlip]],
    checker_name: str,
    db: Session,
) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "考勤记录"

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    title_cell = ws.cell(row=1, column=1, value="吉利学院学风督查记录表（二级学院、辅导员用表）")
    title_cell.font = Font(name="微软雅黑", size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    ws.row_dimensions[1].height = 30

    checker_cell = ws.cell(row=2, column=1, value=f"检查人员：{checker_name}")
    checker_cell.font = Font(name="微软雅黑", size=11)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))

    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    header_fill = __import__("openpyxl.styles", fromlist=["PatternFill"]).PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = COL_WIDTHS[col - 1]

    data_font = Font(name="微软雅黑", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, r in enumerate(records, 1):
        row_idx = i + 3
        college = _get_college(db, r.class_name)
        period = normalize_period(r.period)
        week_label = f"第{r.week_number}周" if r.week_number else ""

        row_data = [
            i, college, r.class_name, _format_export_date(r.date_str), week_label, period,
            r.classroom, r.course_name, r.teacher, r.checker,
            r.class_size, r.actual_count, r.sick_leave_count, r.personal_leave_count,
            r.late_count, r.early_leave_count, r.absent_count, r.not_arrived_count,
            r.attendance_rate, r.attendance_rate_with_leave,
            r.leave_details or "",
            "", "", "", "",
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border

        slips = leave_slips_map.get(r.id, [])
        if slips:
            img_start_col = len(HEADERS) - 3
            for j, slip in enumerate(slips[:4]):
                col_letter = get_column_letter(img_start_col + j)
                _embed_image(ws, slip.image_path, f"{col_letter}{row_idx}", row_idx)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
