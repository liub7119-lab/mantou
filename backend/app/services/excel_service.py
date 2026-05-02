"""
Excel 导出服务
按 4 个成果类别生成不同表头的 Excel 文件，严格对应原始统计表格式
"""

from __future__ import annotations

import tempfile
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from ..config import settings

IMG_WIDTH_PX = 120
IMG_HEIGHT_PX = 90
IMG_ROW_HEIGHT = 72
IMG_COL_WIDTH = 18

# ──────────────────────────────────────────────
# 各类别的表头定义（严格对应 Excel 模板）
# ──────────────────────────────────────────────
CATEGORY_HEADERS = {
    "科研项目": {
        "title": "学生主持、参与教师科研项目情况统计表",
        "headers": [
            "序号", "学号", "学生姓名", "所在学院", "年级专业班级",
            "立项编号（若有）", "项目层级（国家级、省级、市厅级、校级）",
            "参与科研项目名称", "学生排名", "立项年度",
            "项目来源", "项目负责人", "工号", "辅导员姓名",
        ],
        "widths": [6, 15, 10, 22, 22, 18, 18, 35, 10, 10, 18, 12, 15, 12],
        "field_map": lambda idx, student, record: {
            0: idx,
            1: student.student_id,
            2: student.name,
            3: student.major,
            4: student.class_name,
            5: record.project_number,
            6: record.project_level,
            7: record.project_name,
            8: record.student_rank,
            9: record.project_year,
            10: record.project_source,
            11: record.project_leader,
            12: record.leader_id,
            13: record.counselor_name,
        },
    },
    "专利软著": {
        "title": "学生专利（软件著作权）授权情况统计表",
        "headers": [
            "序号", "学号", "学生姓名", "学院", "年级专业班级",
            "名称", "类别", "授权号", "获批时间", "发明人排名", "辅导员姓名",
        ],
        "widths": [6, 15, 10, 22, 22, 35, 15, 22, 14, 12, 12],
        "field_map": lambda idx, student, record: {
            0: idx,
            1: student.student_id,
            2: student.name,
            3: student.major,
            4: student.class_name,
            5: record.name,
            6: record.patent_type,
            7: record.authorization_number,
            8: record.approval_date,
            9: record.inventor_rank,
            10: record.counselor_name,
        },
    },
    "学术论文": {
        "title": "学生发表学术论文情况统计表",
        "headers": [
            "序号", "学号", "学生姓名", "所在学院", "年级专业班级",
            "论文名称", "发表期刊名称", "作者排名", "发表时间",
            "收录情况", "论文链接（若有）", "辅导员姓名",
        ],
        "widths": [6, 15, 10, 22, 22, 35, 25, 10, 14, 12, 30, 12],
        "field_map": lambda idx, student, record: {
            0: idx,
            1: student.student_id,
            2: student.name,
            3: student.major,
            4: student.class_name,
            5: record.paper_name,
            6: record.journal_name,
            7: record.author_rank,
            8: record.publish_date,
            9: record.indexing,
            10: record.paper_link,
            11: record.counselor_name,
        },
    },
    "学科竞赛": {
        "title": "学科竞赛院级及以上获奖情况统计表",
        "headers": [
            "序号", "赛事名称\n(请规范填写赛事全称）", "赛道", "赛事主办方",
            "所在学院", "项目名称", "奖项名称", "项目负责人",
            "学号", "联系电话", "指导老师（全部）",
            "学生成员\n（全部成员，个人赛事则填无）",
            "获奖时间\n（以获奖证书为准，精确到获奖日期）",
            "获奖证书", "备注",
        ],
        "widths": [6, 35, 15, 20, 22, 25, 15, 12, 15, 14, 20, 25, 18, IMG_COL_WIDTH, 20],
        "field_map": lambda idx, student, record: {
            0: idx,
            1: record.competition_name,
            2: record.track,
            3: record.organizer,
            4: student.major,
            5: record.project_name,
            6: record.award_name,
            7: record.team_leader,
            8: student.student_id,
            9: record.phone,
            10: record.advisor,
            11: record.team_members,
            12: record.award_date,
            13: None,  # 证书图片列，单独处理
            14: record.remark,
        },
    },
}


def _apply_styles(ws, headers, widths, title: str):
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # 标题行
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name="微软雅黑", size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.row_dimensions[1].height = 36

    # 表头行
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    return thin_border


def _embed_image(ws, image_path: str, cell_ref: str, row_idx: int):
    if not image_path:
        return
    filename = image_path.lstrip("/").replace("uploads/", "", 1)
    abs_path = Path(settings.UPLOAD_DIR) / filename
    if not abs_path.exists():
        ws[cell_ref].value = "（文件不存在）"
        return
    try:
        pil_img = PILImage.open(abs_path)
        if pil_img.mode in ("RGBA", "P", "LA"):
            pil_img = pil_img.convert("RGB")
        buf = BytesIO()
        pil_img.save(buf, format="JPEG", quality=85)
        buf.seek(0)
        img = XlImage(buf)
        img.width = IMG_WIDTH_PX
        img.height = IMG_HEIGHT_PX
        ws.add_image(img, cell_ref)
    except Exception:
        ws[cell_ref].value = "（图片加载失败）"


def generate_category_excel(category: str, records_with_students: list) -> str:
    """
    按类别生成 Excel 文��
    records_with_students: [(model_instance, student_instance), ...]
    """
    wb = Workbook()
    ws = wb.active

    config = CATEGORY_HEADERS[category]
    ws.title = category
    headers = config["headers"]
    widths = config["widths"]
    field_map_fn = config["field_map"]

    thin_border = _apply_styles(ws, headers, widths, config["title"])

    data_font = Font(name="微软雅黑", size=10)
    data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    has_cert_col = "获奖证书" in headers
    cert_col_idx = headers.index("获奖证书") + 1 if has_cert_col else -1

    for row_offset, (record, student) in enumerate(records_with_students):
        row_idx = row_offset + 3  # 标题行=1, 表头行=2, 数据从3开始
        idx = row_offset + 1
        field_map = field_map_fn(idx, student, record)

        for col_idx_0, value in field_map.items():
            col = col_idx_0 + 1
            if has_cert_col and col == cert_col_idx:
                continue
            cell = ws.cell(row=row_idx, column=col, value=value or "")
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

        if has_cert_col:
            cert_cell = ws.cell(row=row_idx, column=cert_col_idx)
            cert_cell.border = thin_border
            cert_cell.alignment = data_alignment
            ws.row_dimensions[row_idx].height = IMG_ROW_HEIGHT
            cert_ref = f"{get_column_letter(cert_col_idx)}{row_idx}"
            _embed_image(ws, record.certificate_image, cert_ref, row_idx)

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return tmp.name
